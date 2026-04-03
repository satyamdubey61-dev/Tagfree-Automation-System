import tkinter as tk
from tkinter import messagebox, ttk, simpledialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
from tkcalendar import DateEntry
import csv
import os
from history import open_pcba_history_dashboard 

USER_DB_FILE = "users.csv"
PRODUCTS_FILE = "products.csv"

SERVER_EXCEL_PATH = r"\\ingoigw111dat\EA\040_Manufacturing\010_production\060_Repair\pcb\PCBA Debug Analysis\Pcba Rejection/PcbaRejection.xlsx"

def load_products():
    if not os.path.exists(PRODUCTS_FILE):
        return ["C53207-"]
    with open(PRODUCTS_FILE, newline="") as f:
        return [row[0] for row in csv.reader(f) if row]

def save_products(products):
    with open(PRODUCTS_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        for p in products:
            writer.writerow([p])

def create_user_db_file():
    if not os.path.exists(USER_DB_FILE):
        with open(USER_DB_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['username', 'password'])

def add_new_user(username, password):
    users = list_existing_users()
    if username in users:
        return False
    with open(USER_DB_FILE, 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([username, password])
    return True

def check_credentials(username, password):
    with open(USER_DB_FILE, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['username'] == username and row['password'] == password:
                return True
    return False

def list_existing_users():
    if not os.path.exists(USER_DB_FILE):
        return []
    with open(USER_DB_FILE) as f:
        reader = csv.DictReader(f)
        return [row['username'] for row in reader]

def change_password(username, new_password):
    users = []
    with open(USER_DB_FILE, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['username'] == username:
                users.append({'username': username, 'password': new_password})
            else:
                users.append(row)
    with open(USER_DB_FILE, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['username', 'password'])
        writer.writeheader()
        writer.writerows(users)

def delete_account(username, password):
    user_found, passwd_ok = False, False
    users = []
    with open(USER_DB_FILE, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['username'] == username:
                user_found = True
                if row['password'] == password:
                    passwd_ok = True
                    continue
                else:
                    users.append(row)
            else:
                users.append(row)
    if not user_found:
        return "not_found"
    if not passwd_ok:
        return "wrong_password"
    with open(USER_DB_FILE, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['username', 'password'])
        writer.writeheader()
        writer.writerows(users)

    if os.path.exists(SERVER_EXCEL_PATH):
        wb = load_workbook(SERVER_EXCEL_PATH)
        ws = wb.active
        rows_to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[4] == username:
                rows_to_delete.append(idx)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
        wb.save(SERVER_EXCEL_PATH)
    return "deleted"

def startup_chooser():
    chooser = tk.Tk()
    chooser.title("Welcome")
    chooser.geometry("350x180")
    chooser.resizable(False, False)
    tk.Label(chooser, text="Welcome to Pcba Rejection Log", font=('Arial', 14)).pack(pady=(15, 7))
    tk.Label(chooser, text="Select:").pack()
    btn_frame = tk.Frame(chooser)
    btn_frame.pack(pady=10)
    def go_new_user():
        chooser.destroy()
        new_user_window()
    def go_old_user():
        chooser.destroy()
        old_user_window()
    tk.Button(btn_frame, text="New User", width=15, height=2, command=go_new_user, bg="#d0f0c0").grid(row=0, column=0, padx=8)
    tk.Button(btn_frame, text="Old User", width=15, height=2, command=go_old_user, bg="#cce0ff").grid(row=0, column=1, padx=8)
    chooser.mainloop()

def new_user_window():
    win = tk.Tk()
    win.title("Create New User")
    win.geometry("340x210")
    win.resizable(False, False)
    tk.Label(win, text="Create a New User", font=('Arial', 13)).pack(pady=8)
    tk.Label(win, text="Enter Name/ID:").pack(pady=2)
    user_entry = tk.Entry(win, width=28)
    user_entry.pack(pady=2)
    tk.Label(win, text="Enter Password:").pack(pady=2)
    pass_entry = tk.Entry(win, show="*", width=28)
    pass_entry.pack(pady=2)

    def submit_new_user():
        username = user_entry.get().strip()
        password = pass_entry.get().strip()
        if not username or not password:
            messagebox.showerror("Input Error", "Both fields required.", parent=win)
            return
        if ',' in username or ',' in password:
            messagebox.showerror("Invalid Input", "No commas allowed in username/password.")
            return
        if add_new_user(username, password):
            messagebox.showinfo("Success", "Account created! Please use 'Old User' to log in.", parent=win)
            win.destroy()
            old_user_window()
        else:
            messagebox.showerror("User Exists", "That username is taken.", parent=win)
    def do_back():
        win.destroy()
        startup_chooser()
    button_frame = tk.Frame(win)
    button_frame.pack(pady=11)
    tk.Button(button_frame, text="Create", width=13, command=submit_new_user).pack(side=tk.LEFT, padx=7)
    tk.Button(button_frame, text="Back", width=13, command=do_back).pack(side=tk.LEFT, padx=7)
    win.mainloop()

def old_user_window():
    win = tk.Tk()
    win.title("Old User Login")
    win.geometry("370x310")
    win.resizable(False, False)
    users = list_existing_users()
    tk.Label(win, text="Select Your Name", font=('Arial', 12)).pack(pady=6)
    user_var = tk.StringVar()
    user_combo = ttk.Combobox(win, textvariable=user_var, state="readonly", values=users, width=23)
    user_combo.pack(pady=8)
    tk.Label(win, text="Enter Password:").pack()
    pass_entry = tk.Entry(win, show="*", width=25)
    pass_entry.pack(pady=4)

    def try_login():
        user = user_var.get()
        pwd = pass_entry.get()
        if not user or not pwd:
            messagebox.showerror("Error", "Select name and enter password.", parent=win)
            return
        if check_credentials(user, pwd):
            win.destroy()
            launch_main_app(user)
        else:
            messagebox.showerror("Login Failed", "Incorrect name or password.", parent=win)

    def forgot_password():
        username = simpledialog.askstring("Forgot Password", "Enter your username:", parent=win)
        if not username:
            return
        if username not in list_existing_users():
            messagebox.showerror("Not Found", "Username not found!", parent=win)
            return
        while True:
            new_password = simpledialog.askstring("Reset Password", "Enter new password:", parent=win, show="*")
            if new_password is None:
                return
            confirm_password = simpledialog.askstring("Reset Password", "Confirm new password:", parent=win, show="*")
            if confirm_password is None:
                return
            if not new_password.strip():
                messagebox.showerror("Error", "Password cannot be empty.", parent=win)
                continue
            if new_password != confirm_password:
                messagebox.showerror("Error", "Passwords do not match.", parent=win)
                continue
            if ',' in new_password:
                messagebox.showerror("Invalid Input", "No commas allowed in password.", parent=win)
                continue
            break
        change_password(username, new_password)
        messagebox.showinfo("Success", "Password changed successfully! Please log in.", parent=win)

    def do_back():
        win.destroy()
        startup_chooser()

    tk.Button(win, text="Forgot Password?", width=16, command=forgot_password, bg="#f0e68c").pack(pady=(15, 5))
    tk.Button(win, text="Login", width=12, command=try_login).pack(pady=5)
    tk.Button(win, text="Back", width=10, command=do_back, bg="#cccccc").pack(pady=5)
    win.mainloop()

def launch_main_app(username):
    if not os.path.exists(SERVER_EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Date", "Product", "Sr.No", "Defect Observed", "Observed By"])
        wb.save(SERVER_EXCEL_PATH)

    root = tk.Tk()
    root.title("PcbaRejection.xlsx")
    root.geometry("880x650")
    topbar = tk.Frame(root, bg="#eaecec")
    topbar.pack(fill="x", pady=2)
    tk.Label(topbar, text=f"Logged in as: {username}", fg="navy", bg="#eaecec", font=('Arial', 11, 'bold')).pack(side="left", padx=15, pady=3)

    def do_logout():
        root.destroy()
        startup_chooser()

    tk.Button(topbar, text="Log Out", fg="white", bg="#b22222", command=do_logout, width=7).pack(side="right", padx=10, pady=3)

    main_frame = tk.Frame(root)
    main_frame.pack(padx=10, pady=9, fill="both", expand=True)
    labels = ["Date:", "Product:", "Sr.No:", "Defect Observed:", "Observed By:"]
    entries = []
    for idx, label_text in enumerate(labels):
        tk.Label(main_frame, text=label_text, anchor='w').grid(row=idx, column=0, padx=10, pady=5, sticky='w')
        if idx == 0:
            entry = DateEntry(main_frame, width=47, date_pattern='yyyy-mm-dd')
        elif idx == 1:
            product_values = load_products()
            product_var = tk.StringVar()
            entry = ttk.Combobox(main_frame, textvariable=product_var, values=product_values, width=47)
            entry.set(product_values[0] if product_values else "C53207-")
            product_cb = entry
            def add_product():
                new_code = simpledialog.askstring("Add Product", "Enter new product number:", initialvalue="C53207-", parent=root)
                if new_code is None:
                    return
                new_code = new_code.strip()
                if not new_code:
                    messagebox.showerror("Input Error", "Product number cannot be empty.", parent=root)
                    return
                products = load_products()
                if new_code in products:
                    messagebox.showinfo("Exists", "Product already in the list.", parent=root)
                    product_cb.set(new_code)
                    product_cb['values'] = products
                    root.update_idletasks()
                    return
                products.append(new_code)
                save_products(products)
                product_cb['values'] = products
                product_cb.set(new_code)
                root.update_idletasks()
            def remove_product():
                selected_product = product_cb.get().strip()
                if not selected_product:
                    messagebox.showerror("Error", "Select a Product to remove.", parent=root)
                    return
                products = load_products()
                if selected_product not in products:
                    messagebox.showinfo("Not Found", "Product not found in list.", parent=root)
                    return
                products.remove(selected_product)
                save_products(products)
                product_cb['values'] = products
                if products:
                    product_cb.set(products[0])
                else:
                    product_cb.set("")
                root.update_idletasks()
            btn_frame = tk.Frame(main_frame)
            btn_frame.grid(row=idx, column=2, padx=4, pady=5, sticky='w')
            tk.Button(btn_frame, text="Add", command=add_product, width=7).pack(side='left', padx=2)
            tk.Button(btn_frame, text="Remove", command=remove_product, width=7).pack(side='left', padx=2)
        elif idx == 4:
            entry = tk.Entry(main_frame, width=50)
            entry.insert(0, username)
            entry.config(state="readonly")
        else:
            entry = tk.Entry(main_frame, width=50)
        entry.grid(row=idx, column=1, padx=10, pady=5, sticky='ew')
        entries.append(entry)

    date_entry, product_combobox, sr_no_entry, defect_entry, observed_by_entry = entries

    cols = ["Date", "Product", "Sr.No", "Defect Observed", "Observed By"]
    table = ttk.Treeview(main_frame, columns=cols, show="headings", selectmode="browse")
    for col in cols:
        table.heading(col, text=col)
        table.column(col, width=150)
    table.grid(row=6, column=0, columnspan=3, padx=10, pady=18, sticky='nsew')
    main_frame.grid_columnconfigure(1, weight=1)
    main_frame.grid_rowconfigure(6, weight=1)

    # MODIFIED: Only show rows belonging to current user
    def update_table():
        table.delete(*table.get_children())
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            # Only insert rows where Observed By == username
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[4] == username:
                    table.insert("", "end", values=row)
        except Exception as e:
            print("Failed to load Excel data:", e)

    def clear_fields():
        date_entry.set_date(datetime.today())
        if product_combobox['values']:
            product_combobox.set(product_combobox['values'][0])
        sr_no_entry.delete(0, tk.END)
        defect_entry.delete(0, tk.END)
        table.selection_remove(table.focus())

    def get_row_data():
        date = date_entry.get()
        product = product_combobox.get()
        sr_no = sr_no_entry.get()
        defect = defect_entry.get()
        observed_by = username
        if not all([date, product, sr_no, defect, observed_by]):
            messagebox.showerror("Error", "All fields are required!")
            return None
        return [date, product, sr_no, defect, observed_by]

    def save_data():
        row_data = get_row_data()
        if not row_data:
            return
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            ws.append(row_data)
            wb.save(SERVER_EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to save to Excel:\n{e}")
            return
        update_table()
        clear_fields()
        messagebox.showinfo("Success", "Data saved!")

    def load_selected_row(event):
        selected_item = table.focus()
        if selected_item:
            values = table.item(selected_item, 'values')
            clear_fields()
            date_entry.set_date(values[0])
            product_combobox.set(values[1])
            sr_no_entry.delete(0, tk.END)
            sr_no_entry.insert(0, values[2])
            defect_entry.delete(0, tk.END)
            defect_entry.insert(0, values[3])

    def update_selected():
        selected_item = table.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a row to update.")
            return
        row_data = get_row_data()
        if not row_data:
            return
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            excel_row_idx = None
            selected_values = table.item(selected_item, 'values')
            # Find the row for THIS USER only
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if list(row) == list(selected_values):
                    excel_row_idx = idx
                    break
            if not excel_row_idx:
                excel_row_idx = int(table.index(selected_item)) + 2
        except Exception:
            excel_row_idx = int(table.index(selected_item)) + 2
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=excel_row_idx, column=col_idx, value=value)
            wb.save(SERVER_EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to update Excel:\n{e}")
            return
        update_table()
        clear_fields()
        messagebox.showinfo("Success", "Row updated!")

    def delete_selected():
        selected_item = table.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a row to delete.")
            return
        if not messagebox.askyesno("Confirm Delete", "Delete selected row?"):
            return
        selected_values = table.item(selected_item, 'values')
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            excel_row_idx = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if list(row) == list(selected_values):
                    excel_row_idx = idx
                    break
            if not excel_row_idx:
                excel_row_idx = int(table.index(selected_item)) + 2
        except Exception:
            excel_row_idx = int(table.index(selected_item)) + 2
        try:
            wb = load_workbook(SERVER_EXCEL_PATH)
            ws = wb.active
            ws.delete_rows(excel_row_idx)
            wb.save(SERVER_EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to delete from Excel:\n{e}")
            return
        update_table()
        clear_fields()

    table.bind("<<TreeviewSelect>>", load_selected_row)

    button_frame = tk.Frame(main_frame)
    button_frame.grid(row=5, column=0, columnspan=3, pady=7)
    tk.Button(button_frame, text="Save", command=save_data, bg="green", fg="white", width=15).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Update Selected", command=update_selected, bg="blue", fg="white", width=15).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Delete Selected", command=delete_selected, bg="red", fg="white", width=15).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Pcba History", command=lambda: open_pcba_history_dashboard(root), bg="#ffcc00", fg="black", width=15).pack(side=tk.LEFT, padx=5)

    update_table()
    root.mainloop()

if __name__ == "__main__":
    create_user_db_file()
    startup_chooser()