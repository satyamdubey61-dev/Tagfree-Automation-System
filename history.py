import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

def open_pcba_history_dashboard(parent_win=None):
    dashboard = tk.Toplevel(parent_win)
    dashboard.title("PCBA History Dashboard")
    dashboard.geometry("1050x680")
    dashboard.resizable(True, True)

    # Serial number entry and scan button (side by side)
    top_frame = tk.Frame(dashboard)
    top_frame.pack(pady=16)
    tk.Label(top_frame, text="Enter Serial Number:", font=('Arial', 13)).pack(side=tk.LEFT, padx=(2,8))
    serial_entry = tk.Entry(top_frame, width=36, font=('Arial', 12))
    serial_entry.pack(side=tk.LEFT, padx=(0,11))
    # Scan button
    tk.Button(top_frame, text="Scan", command=lambda: scan_pcba_history(), bg="#1976d2", fg="white", font=('Arial', 12), width=14, height=1).pack(side=tk.LEFT)

    # List to hold search sources
    source_paths = []

    # Sources section
    src_frame = tk.LabelFrame(dashboard, text="Add Folder or Excel Source", font=('Arial', 12), padx=10, pady=8)
    src_frame.pack(fill='x', padx=26, pady=(10,8))

    def add_folder():
        folder = filedialog.askdirectory(title='Select Server Directory')
        if folder and folder not in source_paths:
            source_paths.append(folder)
            update_source_list()

    def add_excel():
        excel_file = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx;*.xls')])
        if excel_file and excel_file not in source_paths:
            source_paths.append(excel_file)
            update_source_list()

    tk.Button(src_frame, text="Add Folder", command=add_folder, width=18, font=('Arial', 11)).pack(side=tk.LEFT, padx=6)
    tk.Button(src_frame, text="Add Excel Sheet", command=add_excel, width=18, font=('Arial', 11)).pack(side=tk.LEFT, padx=6)

    # Source list display
    src_list_frame = tk.Frame(dashboard)
    src_list_frame.pack(fill='x', padx=26, pady=(3,10))
    src_list_label = tk.Label(src_list_frame, text="Search Sources:", font=('Arial', 11, 'bold'))
    src_list_label.pack(anchor='w')
    src_listbox = tk.Listbox(src_list_frame, width=110, height=3, font=('Arial', 10))
    src_listbox.pack(side=tk.LEFT, padx=(0,10))
    def remove_selected_source():
        selected = src_listbox.curselection()
        if selected:
            idx = selected[0]
            removed = source_paths.pop(idx)
            update_source_list()
    tk.Button(src_list_frame, text="Remove Selected", command=remove_selected_source, width=22, font=('Arial', 11)).pack(side=tk.LEFT)

    def update_source_list():
        src_listbox.delete(0, tk.END)
        for path in source_paths:
            src_listbox.insert(tk.END, path)

    # Results box
    result_box = scrolledtext.ScrolledText(dashboard, width=130, height=24, font=('Consolas', 11))
    result_box.pack(padx=10, pady=12, fill='both', expand=True)

    # Scan logic
    def scan_pcba_history():
        serial = serial_entry.get().strip()
        result_box.delete("1.0", tk.END)
        if not serial:
            messagebox.showerror("Missing Data", "Please enter a serial number.")
            return
        if not source_paths:
            messagebox.showerror("Missing Data", "Please add at least one source folder or file.")
            return

        result_text = f"PCBA History for Serial Number: {serial}\n" + "-"*104 + "\n\n"
        any_found = False
        # Loop through each source
        for source in source_paths:
            if os.path.isdir(source):
                for root_dir, dirs, files in os.walk(source):
                    for file in files:
                        full_path = os.path.join(root_dir, file)
                        try:
                            # TXT files (line-wise exact match for serial)
                            if file.lower().endswith('.txt'):
                                with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    matched_lines = []
                                    for line in f:
                                        # Split line by whitespace or , for cells, then check for exact serial
                                        cells = [cell.strip() for cell in line.replace(',', ' ').split()]
                                        if serial in cells:
                                            matched_lines.append(line.rstrip())
                                    if matched_lines:
                                        any_found = True
                                        result_text += f"[TXT] {full_path}:\n"
                                        for mline in matched_lines:
                                            result_text += mline + '\n'
                                        result_text += '\n'
                            # XML files (element text/content exact match for serial)
                            elif file.lower().endswith('.xml'):
                                try:
                                    tree = ET.parse(full_path)
                                    xml_root = tree.getroot()
                                    found_this_file = False
                                    for elem in xml_root.iter():
                                        # Check element tag or text for exact serial
                                        if elem.text and elem.text.strip() == serial:
                                            if not found_this_file:
                                                result_text += f"[XML] {full_path}:\n"
                                                found_this_file = True
                                            result_text += ET.tostring(elem, encoding='unicode') + '\n'
                                    if found_this_file:
                                        any_found = True
                                        result_text += '\n'
                                except ET.ParseError:
                                    pass
                        except Exception as e:
                            result_text += f"Error reading {full_path}: {e}\n"
            # Excel file, search for exact cell matches only
            elif os.path.isfile(source) and source.lower().endswith(('.xlsx', '.xls')):
                try:
                    wb = load_workbook(source)
                    ws = wb.active
                    found_excel = False
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if str(cell).strip() == serial:
                                row_data = ", ".join(str(c) for c in row)
                                if not found_excel:
                                    result_text += f"[Excel] ({os.path.basename(source)}):\n"
                                    found_excel = True
                                result_text += row_data + "\n"
                                any_found = True
                    if found_excel:
                        result_text += '\n'
                except Exception as e:
                    result_text += f"Excel error in {os.path.basename(source)}: {e}\n"
            else:
                result_text += f"Unsupported source type: {source}\n"
        if not any_found:
            result_box.insert(tk.END, "No history found for serial number: " + serial)
        else:
            result_box.insert(tk.END, result_text)
